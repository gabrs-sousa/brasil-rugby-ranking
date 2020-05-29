import pandas as pd
from openpyxl import worksheet


def map_team_names(names_sheet: worksheet, games_sheet: worksheet):
    mapped_names = []
    missing_names = set()

    names_last_row = names_sheet.max_row
    for row in range(1, names_last_row):
        team_name = names_sheet.cell(row, 1).value

        if team_name:
            mapped_names.append(team_name.upper())

    games_last_row = games_sheet.max_row

    for row in range(2, games_last_row):
        visitor = games_sheet.cell(row, 7).value
        home = games_sheet.cell(row, 12).value

        if home and home.upper() not in mapped_names:
            missing_names.add(home)

        if visitor and visitor.upper() not in mapped_names:
            missing_names.add(visitor)

    if missing_names:
        return missing_names
    else:
        return False


def format_name(name: str) -> str:
    """
    Limpa espaços antes e depois da palavra
    Nome em caps lock para evitar case sensitive
    """
    name = name.strip() 
    name = name.upper() 
    return name


def export_output_file(teams: dict, output_file_name: str):
    ranking_df = pd.DataFrame(teams)
    ranking_df = ranking_df.transpose()
    ranking_df = ranking_df.sort_values('points', ascending=False)
    ranking_df = ranking_df[ranking_df['total_games'] > 0].dropna()
    ranking_df.to_excel(output_file_name)
    print(f'Workbook "{output_file_name}" has been created successfully!')
