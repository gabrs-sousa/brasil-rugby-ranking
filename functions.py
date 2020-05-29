from datetime import date

from openpyxl import worksheet

from utils import format_name


# Regra 1: Home team começa com (pontuação atual + 3)
# Regra 2: Calcular diferença de pontuações Home/Visitor (Rating Gap)
# Regra 3: [Points_Exchange] = -(0.1 * [Team_Difference] - [Result])
# Regra 4: Maximum Team_Difference locks at 10
# Regra 5: Result é +1 quando Home ganha, 0 quando empata e -1 quando Home perde
# Se for campeonato Brasileiro, multiplica resultado por 2

# Example: Going back to the Samoa vs United States 4.78 difference. Samoa won by less than 15.
# -(0.1 * Team_Difference - Result) = Points_Exchange
# -(0.1 * [4.78] - [+1]) = -(0.478 - [+1]) = -(-0.522) = 0.522


def find_official_name(unofficial_name: str, teams: dict):
    for team in teams.keys():
        name_variations = teams[team]['variations']

        if unofficial_name in name_variations:
            return team
    
    # Vai vir pra cá caso o nome não seja encontrado
    return None


def format_and_map_games_data(games_sheet: worksheet, teams_dict: dict) -> dict:
    games = {}
    last_row_games = games_sheet.max_row

    for row in range(2, last_row_games):
        game_id = games_sheet.cell(row, 1).value

        home_team = games_sheet.cell(row, 7).value
        home_score = games_sheet.cell(row, 9).value

        away_team = games_sheet.cell(row, 12).value
        away_score = games_sheet.cell(row, 11).value

        game_date = games_sheet.cell(row, 2).value
        championship_name = format_name(games_sheet.cell(row, 15).value) if games_sheet.cell(row, 15).value else None

        if not away_team or not home_team:
            print(f'JOGO NÃO MAPEADO: O jogo "{game_id}" não tem os nomes das duas equipes.')
            continue
        
        home_team = format_name(home_team)
        away_team = format_name(away_team)
        
        if home_team not in teams_dict.keys():
            this_name = away_team
            home_team = find_official_name(home_team, teams_dict)
            
            if not home_team:
                print(f'JOGO NÃO MAPEADO: O time "{this_name}" não está na lista de nomes.')
                continue
        
        if away_team not in teams_dict.keys():
            this_name = away_team
            away_team = find_official_name(away_team, teams_dict)

            if not away_team:
                print(f'JOGO NÃO MAPEADO: O time "{this_name}" não está na lista de nomes.')
                continue

        if str(home_score) == 'None' or str(away_score) == 'None':
            print(f'JOGO NÃO MAPEADO: O jogo "{game_id}" não tem pontuação das duas equipes.')
            continue
        
        if not game_date or not isinstance(game_date, date):
            # Pula caso o jogo não tenha data ou data não esteja bem formatada
            print(f'JOGO NÃO MAPEADO: O jogo "{game_id}" não tem data ou a data está mal formatada.')
            continue
        
        if home_score > away_score:
            winner = 'home'
        
        elif home_score < away_score:
            winner = 'away'
        
        else:
            winner = 'draw'
        
        games[game_id] = {
            'game_date': game_date,
            'game_time': games_sheet.cell(row, 3).value,
            'game_class': games_sheet.cell(row, 4).value,
            'genre': games_sheet.cell(row, 5).value,
            'home_team': home_team,
            'home_team_state': games_sheet.cell(row, 8).value,
            'home_score': home_score,
            'away_score': away_score,
            'away_team': away_team,
            'away_team_state': games_sheet.cell(row, 13).value,
            'game_location': games_sheet.cell(row, 14).value,
            'championship': championship_name,
            'game_city':games_sheet.cell(row, 16).value,
            'game_state': games_sheet.cell(row, 17).value,
            
            # Campos calculados
            'row': row,
            'winner': winner,
            'double_points': True if championship_name and 'CAMPEONATO BRASILEIRO - SÉRIE A' in championship_name else False,
            '15+': True if (home_score - away_score) >= 15 or (home_score - away_score) <= -15 else False,
        }

    return games


def map_team_names(name_sheet: worksheet) -> dict:
    """
    Mapeia os nomes dos times e suas variações
    """
    row = 0
    names = {}
    last_row = name_sheet.max_row

    while True:
        row += 1
        next_row = row + 1
        team = name_sheet.cell(row, 1).value
        variation = name_sheet.cell(next_row, 1).value

        if row > last_row:
            # Para quando chega à última linha
            break

        if not team:
            continue

        team = format_name(team)

        names[team] = {
            'wins': 0,
            'losses': 0,
            'draws': 0,
            'total_games': 0,
            'points': 0,
            'variations': []
        }

        while variation:
            variation = format_name(variation)
            if variation not in names[team]['variations']:
                names[team]['variations'].append(variation)

            next_row += 1
            variation = name_sheet.cell(next_row, 1).value
            row = next_row

    return names


def calculate_point_exchange(game: dict, teams: dict):
    home_team_name = game['home_team']
    away_team_name = game['away_team']

    home_team = teams[home_team_name]
    away_team = teams[away_team_name]

    # +3 related to home advantage
    rating_gap = (home_team['points'] + 3) - away_team['points']

    if rating_gap > 10:
        rating_gap = 10

    elif rating_gap < -10:
        rating_gap = -10

    if game['winner'] == 'home':
        # Home team wins
        home_team['wins'] += 1
        away_team['losses'] += 1
        exchange_rate = -(0.1 * rating_gap - 1)

    elif game['winner'] == 'away':
        away_team['wins'] += 1
        home_team['losses'] += 1
        exchange_rate = -(0.1 * rating_gap + 1)

    else:
        home_team['draws'] += 1
        away_team['draws'] += 1
        exchange_rate = -(0.1 * rating_gap - 0)

    if game['15+']:
        exchange_rate *= 1.5

    if game['double_points']:
        exchange_rate *= 2

    home_team['total_games'] += 1
    away_team['total_games'] += 1

    home_team['points'] += exchange_rate
    away_team['points'] -= exchange_rate


def calculate_win_rate(teams):
    for team_details in teams.values():
        
        if not team_details['total_games'] \
                or team_details['total_games'] < 10:

            team_details['points'] = 0
            continue

        win_rate = (team_details['wins'] / team_details['total_games']) * 100
        team_details['points'] = win_rate


def set_initial_points(teams: dict, points: int):
    for team_details in teams.values():
        team_details['points'] = points
    

def calculate_scores(teams: dict, games: dict, calculate_pre_score: bool):
    if calculate_pre_score:
        for game in games.keys():
            this_game = games[game]

            if this_game['game_date'].year < date.today().year:
                calculate_point_exchange(this_game, teams)
                
        calculate_win_rate(teams)
    
    elif not calculate_pre_score:
        set_initial_points(teams, 40)
    
    # Vem pra cá após calcular
    # Os pontos iniciais
    for game in games.keys():
        this_game = games[game]
        calculate_point_exchange(this_game, teams)
